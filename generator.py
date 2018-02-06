import os
from copy import copy
import six

from openpyxl import formatting
from openpyxl import styles
from openpyxl import Workbook
from openpyxl.cell.cell import get_column_letter

from utils import tLinkType

class TraceabilityGenerator:
    ''' Utility class for generating a requirements traceability matrix'''
    
    HEADER_FONT = styles.Font(
            name='Calibri',
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
    CELL_FONT = styles.Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
    SUMMARY_FONT = styles.Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
    HEADER_ALIGNMENT = styles.Alignment(
            horizontal='center',
            vertical='bottom',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0)
    CELL_ALIGNMENT = styles.Alignment(
            horizontal='left',
            vertical='bottom',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0)
    SUMMARY_ALIGNMENT = styles.Alignment(
            horizontal='center',
            vertical='bottom',
            text_rotation=0,
            wrap_text=True,
            shrink_to_fit=False,
            indent=0)
    REQ_NOT_MET_FILL = styles.PatternFill(
            start_color='EE1111', end_color='EE1111', fill_type='solid')
    REQ_MET_FILL = styles.PatternFill(
            start_color='0FFFFF', end_color='0FFFFF', fill_type='solid')
    
    COL_COUNT_FORMULA = 'COUNTA(\'%s\'!%s:%s)-1'
    PERCENT_FORMULA = '=IF(%s%d, %s%d/%s%d, 0.0)'
    REQ_IS_MET_FORMULA = 'OR(%s,ISBLANK($%s$%d))'
                    
    @staticmethod
    def generateTraceabilityMatrix(reqMap, args):
        ''' Generates a workbook with a traceability matrix from a 
        requirements map'''
        
        outputDir = os.path.expanduser(args.outputDir)
        outputDir = os.path.expandvars(outputDir)
        
        # make sub-directories for exported traceability matrix
        if (not os.path.exists(outputDir)):
            os.makedirs(outputDir)
        
        outfile = os.path.join(outputDir, args.outfile + '.xlsx')
        
        wb = Workbook()

        # generate summary sheet
        summarySheet = wb.active
        TraceabilityGenerator._generateTraceabilitySummary(summarySheet, reqMap, args)
        
        # generate sheet for each module
        for moduleName, module in six.iteritems(reqMap):
            TraceabilityGenerator._generateTraceabilitySheet(wb, moduleName, module, args)
        
        wb.save(outfile)
    
    @staticmethod
    def _generateTraceabilitySummary(sheet, reqMap, args):
        ''' Generates a summary sheet with summary details
        for each module and overall summary'''
        
        sheet.title = 'Summary'
        
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 3
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 3
        sheet.column_dimensions['G'].width = 15
        
        cellRow = 1
        for moduleName in reqMap:
            cellRow += TraceabilityGenerator._generateModuleSummary(sheet, moduleName, cellRow, 1, args)
            cellRow += 1
            
        # set summary of module summaries
        cell = sheet.cell(row=cellRow, column=1)
        cell.value = 'Summary'
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
        
        cellRow += 1
        moduleCol = 3
        
        expectedCount = TraceabilityGenerator._getSummaryFormula(reqMap, 'A')
        
        if (True == args.checkSrcLinks):
            # set summary of source links
            actualCount = TraceabilityGenerator._getSummaryFormula(reqMap, get_column_letter(moduleCol))
            TraceabilityGenerator._generateSummaryColSummary(sheet, 'Source Links:', actualCount, expectedCount, cellRow, 2)
            
            cellRow += 1
            moduleCol += 1
        
        if (True == args.checkTestLinks):
            # set summary of test links
            actualCount = TraceabilityGenerator._getSummaryFormula(reqMap, get_column_letter(moduleCol))
            TraceabilityGenerator._generateSummaryColSummary(sheet, 'Test Links:', actualCount, expectedCount, cellRow, 2)
            
            cellRow += 1
            moduleCol += 1
        
    @staticmethod
    def _getSummaryFormula(reqMap, colStr):
        ''' Generates an summary formula for all the modules
        for the specified column (requirement type)'''
        
        countFormula = '=0'
        for moduleName in reqMap:
            countFormula += '+' + TraceabilityGenerator.COL_COUNT_FORMULA % (moduleName, colStr, colStr)
        return countFormula
    
    @staticmethod
    def _generateSummaryColSummary(sheet, colName, countFormula, totalFormula, rowOffset, colOffset):
        ''' Generates a column summary details for all the modules in the format
        <summary_description> | <actual_value> / <expected_value> = <percent_value>'''
        
        # set summary type
        cell = sheet.cell(row=rowOffset, column=colOffset)
        cell.value = colName
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set actual summary count
        cell = sheet.cell(row=rowOffset, column=colOffset+1)
        cell.value = countFormula
        cell.font = TraceabilityGenerator.SUMMARY_FONT
        alignment = copy(TraceabilityGenerator.SUMMARY_ALIGNMENT)
        alignment.horizontal = 'right'
        cell.alignment = alignment
        
        cell = sheet.cell(row=rowOffset, column=colOffset+2)
        cell.value = '/'
        font = copy(TraceabilityGenerator.SUMMARY_FONT)
        font.bold = True
        cell.font = font
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set expected summary count
        cell = sheet.cell(row=rowOffset, column=colOffset+3)
        cell.value = totalFormula
        cell.font = TraceabilityGenerator.SUMMARY_FONT
        alignment = copy(TraceabilityGenerator.SUMMARY_ALIGNMENT)
        alignment.horizontal = 'left'
        cell.alignment = alignment
        
        cell = sheet.cell(row=rowOffset, column=colOffset+4)
        cell.value = '='
        font = copy(TraceabilityGenerator.SUMMARY_FONT)
        font.bold = True
        cell.font = font
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set percentage of actual/expected
        cell = sheet.cell(row=rowOffset, column=(colOffset+5))
        cell.value = TraceabilityGenerator.PERCENT_FORMULA % (\
                get_column_letter(colOffset+3), rowOffset,
                get_column_letter(colOffset+1), rowOffset, 
                get_column_letter(colOffset+3), rowOffset)
        cell.font = TraceabilityGenerator.SUMMARY_FONT
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        cell.number_format = '0.00%'
        
    @staticmethod
    def _generateModuleColSummary(sheet, moduleName, colName, countCol, totalCol, rowOffset, colOffset):
        ''' Generates a column summary details for the specified module in the format
        <summary_description> | <actual_value> / <expected_value> = <percent_value>'''
        
        # set column summary type
        cell = sheet.cell(row=rowOffset, column=colOffset)
        cell.value = colName
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set column actual count
        cell = sheet.cell(row=rowOffset, column=(colOffset+1))
        cell.value = '=' + TraceabilityGenerator.COL_COUNT_FORMULA % (moduleName, get_column_letter(countCol), get_column_letter(countCol))
        alignment = copy(TraceabilityGenerator.SUMMARY_ALIGNMENT)
        alignment.horizontal = 'right'
        cell.alignment = alignment
        
        cell = sheet.cell(row=rowOffset, column=(colOffset+2))
        cell.value = '/'
        font = copy(TraceabilityGenerator.SUMMARY_FONT)
        font.bold = True
        cell.font = font
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set column expected count
        cell = sheet.cell(row=rowOffset, column=(colOffset+3))
        cell.value = '=' + TraceabilityGenerator.COL_COUNT_FORMULA % (moduleName, get_column_letter(totalCol), get_column_letter(totalCol))
        cell.font = TraceabilityGenerator.SUMMARY_FONT
        alignment = copy(TraceabilityGenerator.SUMMARY_ALIGNMENT)
        alignment.horizontal = 'left'
        cell.alignment = alignment
        
        cell = sheet.cell(row=rowOffset, column=(colOffset+4))
        cell.value = '='
        font = copy(TraceabilityGenerator.SUMMARY_FONT)
        font.bold = True
        cell.font = font
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        
        # set actual / expected percentage
        cell = sheet.cell(row=rowOffset, column=(colOffset+5))
        cell.value = TraceabilityGenerator.PERCENT_FORMULA % (\
                get_column_letter(colOffset+3), rowOffset,
                get_column_letter(colOffset+1), rowOffset, 
                get_column_letter(colOffset+3), rowOffset)
        cell.font = TraceabilityGenerator.SUMMARY_FONT
        cell.alignment = TraceabilityGenerator.SUMMARY_ALIGNMENT
        cell.number_format = '0.00%'
        
    @staticmethod
    def _generateModuleSummary(sheet, moduleName, rowOffset, colOffset, args):
        ''' Generates a summary section for the specified module'''
        
        cellRow = 0
        
        # set module name
        cell = sheet.cell(row=rowOffset+cellRow, column=colOffset)
        cell.value = moduleName
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
        
        cellRow += 1
        
        # start at offset because of default columns: (name | text | satisfied)
        moduleCol = 4
        
        if (True == args.checkSrcLinks):
            # set source links summary
            TraceabilityGenerator._generateModuleColSummary(sheet, moduleName, 'Source Links:', moduleCol, 1, rowOffset+cellRow, colOffset+1)
            
            moduleCol += 1
            cellRow += 1
        
        if (True == args.checkTestLinks):
            # set test links summary
            TraceabilityGenerator._generateModuleColSummary(sheet, moduleName, 'Test Links:', moduleCol, 1, rowOffset+cellRow, colOffset+1)
            
            cellRow += 1
        
        return cellRow
        
    @staticmethod
    def _generateTraceabilitySheet(workbook, moduleName, module, args):
        ''' Generates a sheet with all the requirements, requirements 
        details, and requirement links for the specified module'''
        
        # create sheet for module
        moduleSheet = workbook.create_sheet(title=moduleName)
        moduleSheet.fill = TraceabilityGenerator.REQ_MET_FILL
        
        # set column dimensions for requirement name
        moduleSheet.column_dimensions['A'].width = 30
        # set column dimensions for requirement text
        moduleSheet.column_dimensions['B'].width = 70
        # set header and requirement text as fixed
        moduleSheet.freeze_panes = 'B2'
        # set column for requirement satisfaction
        moduleSheet.column_dimensions['C'].hidden = True
        
        cellRow = 1
        
        # setup header
        cellCol = 1
        # set start column for conditional formatting
        startCol = get_column_letter(cellCol)
        
        # add requirement name column
        cell = moduleSheet.cell(row=cellRow, column=cellCol)
        cell.value = 'Requirement Name'
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
        
        cellCol += 1
        
        # add requirement text column
        cell = moduleSheet.cell(row=cellRow, column=cellCol)
        cell.value = 'Requirement Text'
        cell.font = TraceabilityGenerator.HEADER_FONT
        cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
        
        cellCol += 1
        
        areReqLinksChecked = \
            (True == args.checkSrcLinks) or \
            (True == args.checkTestLinks)
                            
        if (True == areReqLinksChecked):
            # add requirement satisfied column
            cell = moduleSheet.cell(row=cellRow, column=cellCol)
            cell.value = 'SATISFIED'
            
            cellCol += 1
            
            if (True == args.checkSrcLinks):
                # set column dimension for test links
                moduleSheet.column_dimensions[get_column_letter(cellCol)].width = 75
                
                cell = moduleSheet.cell(row=cellRow, column=cellCol)
                cell.value = 'Source Code Links'
                cell.font = TraceabilityGenerator.HEADER_FONT
                cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
                
                cellCol += 1
            
            if (True == args.checkTestLinks):
                # set column dimension for test links
                moduleSheet.column_dimensions[get_column_letter(cellCol)].width = 75
                
                cell = moduleSheet.cell(row=cellRow, column=cellCol)
                cell.value = 'Test Links'
                cell.font = TraceabilityGenerator.HEADER_FONT
                cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
                
                cellCol += 1
        
        # get end column for conditional formatting
        endCol = get_column_letter(cellCol - 1)
        
        cellRow += 1
        
        # add row for each requirement in module
        for req, reqValue in six.iteritems(module):
            rowCol = 1
            
            # requirement name
            cell = moduleSheet.cell(row=cellRow, column=rowCol)
            cell.value = req
            cell.font = TraceabilityGenerator.HEADER_FONT
            cell.alignment = TraceabilityGenerator.HEADER_ALIGNMENT
            
            rowCol += 1
            
            # requirement text
            cell = moduleSheet.cell(row=cellRow, column=rowCol)
            cell.value = reqValue.reqText
            cell.font = TraceabilityGenerator.CELL_FONT
            cell.alignment = TraceabilityGenerator.CELL_ALIGNMENT
            
            rowCol += 1
            
            if (True == areReqLinksChecked):
                # requirement satisfied
                reqMetCell = moduleSheet.cell(row=cellRow, column=rowCol)
                reqMetFormula = 'False'
                
                rowCol += 1
                
                # TODO add requirements as hyperlinks
                '=HYPERLINK(<URL>, <text>)'
                
                if (True == args.checkSrcLinks):
                    # add source links
                    linksText = ''
                    for link in reqValue.reqLinks:
                        if (tLinkType.LINK_TYPE__SRC == link.linkType):
                            linksText += link.linkName + '\n'
                    
                    cell = moduleSheet.cell(row=cellRow, column=rowCol)
                    cell.value = linksText
                    cell.font = TraceabilityGenerator.CELL_FONT
                    cell.alignment = TraceabilityGenerator.CELL_ALIGNMENT
                    reqMetFormula = TraceabilityGenerator.REQ_IS_MET_FORMULA % (reqMetFormula, get_column_letter(rowCol), cellRow)
                    
                    rowCol += 1
                
                if (True == args.checkTestLinks):
                    # add test links
                    linksText = ''
                    for link in reqValue.reqLinks:
                        if (tLinkType.LINK_TYPE__TEST == link.linkType):
                            linksText += link.linkName + '\n'
                    
                    cell = moduleSheet.cell(row=cellRow, column=rowCol)
                    cell.value = linksText
                    cell.font = TraceabilityGenerator.CELL_FONT
                    cell.alignment = TraceabilityGenerator.CELL_ALIGNMENT
                    reqMetFormula = TraceabilityGenerator.REQ_IS_MET_FORMULA % (reqMetFormula, get_column_letter(rowCol), cellRow)
                    
                    rowCol += 1
                
                # set requirement met column value
                reqMetCell.value = '=IF(NOT(%s), "PASS", "FAIL")' % (reqMetFormula)
                
                # set conditional formatting for if requirement is met
                moduleSheet.conditional_formatting.add(
                    '%s%d:%s%d' % (startCol, cellRow, endCol, cellRow),
                    formatting.rule.FormulaRule(formula=[reqMetFormula], stopIfTrue=True, fill=TraceabilityGenerator.REQ_NOT_MET_FILL))
                
            cellRow += 1
            